import openpyxl
from dashboard.models import (
    Asset, Commodity, Company, Company_Policy, Company_Revenue,
    Country, Ownership, Policy_Level, Policy_Subcategory, Policy_Type,
    Production, SubnationalRegion,
)
from .constants import (
    DUPLICATE_CRITERIA, FK_FIELDS, MODEL_KEY_TO_SHEET,
    REQUIRED_FIELDS, SHEET_COLUMNS,
)


def parse_file(source):
    """
    Parse an xlsx file (path or file-like object).
    Returns {sheet_name: [{'status': 'ok'|'duplicate'|'error', 'data': {...}, 'message': str}, …]}.
    """
    wb = openpyxl.load_workbook(source)
    file_names = _collect_file_names(wb)
    db_name_cache = _build_db_name_cache()

    result = {}
    for sheet_name in SHEET_COLUMNS:
        if sheet_name not in wb.sheetnames:
            continue
        result[sheet_name] = _parse_sheet(wb[sheet_name], sheet_name, file_names, db_name_cache)
    return result


# ── helpers ──────────────────────────────────────────────────────────────────

def _collect_file_names(wb):
    """
    Build lookup sets of names defined in the file itself, so FK fields can
    reference rows from a sibling sheet in the same upload.
    Returns {model_key: {name_lower, …}}.
    """
    file_names = {key: set() for key in MODEL_KEY_TO_SHEET}
    for model_key, sheet_name in MODEL_KEY_TO_SHEET.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        if 'name' not in header:
            continue
        name_col = header.index('name')
        for row in ws.iter_rows(min_row=2):
            val = row[name_col].value if name_col < len(row) else None
            if val is not None:
                file_names[model_key].add(str(val).strip().lower())
    return file_names


def _build_db_name_cache():
    """
    Fetch all FK-referenced name sets from the DB in one go.
    Returns {model_key: {name_lower, …}}.
    """
    return {
        'country': {n.lower() for n in Country.objects.values_list('name', flat=True)},
        'subnational_region': {n.lower() for n in SubnationalRegion.objects.values_list('name', flat=True)},
        'commodity': {n.lower() for n in Commodity.objects.values_list('name', flat=True)},
        'policy_type': {n.lower() for n in Policy_Type.objects.values_list('name', flat=True)},
        'policy_subcategory': {n.lower() for n in Policy_Subcategory.objects.values_list('name', flat=True)},
        'policy_level': {n.lower() for n in Policy_Level.objects.values_list('name', flat=True)},
        'company': {n.lower() for n in Company.objects.values_list('name', flat=True)},
        'asset': {n.lower() for n in Asset.objects.values_list('name', flat=True)},
    }


def _can_resolve(model_key, name, file_names, db_name_cache):
    key = name.strip().lower()
    return key in db_name_cache.get(model_key, set()) or key in file_names.get(model_key, set())


def _existing_keys(sheet_name):
    """Return the set of existing duplicate-key tuples from the DB."""
    if sheet_name == 'Country':
        return {(n.lower(),) for n in Country.objects.values_list('name', flat=True)}
    if sheet_name == 'SubnationalRegion':
        return {(n.lower(), c.lower()) for n, c in
                SubnationalRegion.objects.values_list('name', 'country__name')}
    if sheet_name == 'Commodity':
        return {(n.lower(),) for n in Commodity.objects.values_list('name', flat=True)}
    if sheet_name == 'Policy_Type':
        return {(n.lower(),) for n in Policy_Type.objects.values_list('name', flat=True)}
    if sheet_name == 'Policy_Subcategory':
        return {(n.lower(), p.lower()) for n, p in
                Policy_Subcategory.objects.values_list('name', 'policy_type__name')}
    if sheet_name == 'Policy_Level':
        return {(n.lower(), s.lower(), p.lower()) for n, s, p in
                Policy_Level.objects.values_list(
                    'name', 'subcategory__name', 'subcategory__policy_type__name')}
    if sheet_name == 'Company':
        return {(n.lower(),) for n in Company.objects.values_list('name', flat=True)}
    if sheet_name == 'Asset':
        return {(n.lower(), c.lower()) for n, c in
                Asset.objects.values_list('name', 'country__name')}
    if sheet_name == 'Production':
        return {((a or '').lower(), c.lower(), str(y)) for a, c, y in
                Production.objects.values_list('asset__name', 'commodity__name', 'year')}
    if sheet_name == 'Company_Revenue':
        return {(c.lower(), str(y)) for c, y in
                Company_Revenue.objects.values_list('company__name', 'year')}
    if sheet_name == 'Ownership':
        return {(a.lower(), c.lower()) for a, c in
                Ownership.objects.values_list('Asset__name', 'Company__name')}
    if sheet_name == 'Company_Policy':
        return {(co.lower(), (pt or '').lower(), (ps or '').lower(), (pl or '').lower())
                for co, pt, ps, pl in Company_Policy.objects.values_list(
                    'company__name',
                    'policy_level__subcategory__policy_type__name',
                    'policy_level__subcategory__name',
                    'policy_level__name',
                )}
    return set()


def _parse_sheet(ws, sheet_name, file_names, db_name_cache):
    columns = SHEET_COLUMNS[sheet_name]
    required = REQUIRED_FIELDS[sheet_name]
    fk_fields = FK_FIELDS.get(sheet_name, {})
    dup_criteria = DUPLICATE_CRITERIA[sheet_name]

    existing = _existing_keys(sheet_name)
    seen = set()
    rows_out = []

    header = [c.value for c in ws[1]]
    header_map = {name: idx for idx, name in enumerate(header) if name is not None}

    for ws_row in ws.iter_rows(min_row=2):
        data = {}
        for col_name in columns:
            col_idx = header_map.get(col_name)
            if col_idx is not None and col_idx < len(ws_row):
                cell_val = ws_row[col_idx].value
            else:
                cell_val = None
            data[col_name] = str(cell_val).strip() if cell_val is not None else ''

        if all(v == '' for v in data.values()):
            continue

        # Required fields
        missing = [f for f in required if not data.get(f, '')]
        if missing:
            rows_out.append({
                'status': 'error',
                'message': f"Champs obligatoires manquants : {', '.join(missing)}",
                'data': data,
            })
            continue

        # FK resolution
        fk_error = None
        for fk_col, model_key in fk_fields.items():
            val = data.get(fk_col, '')
            if val and not _can_resolve(model_key, val, file_names, db_name_cache):
                fk_error = f"Valeur introuvable pour '{fk_col}' : '{val}'"
                break
        if fk_error:
            rows_out.append({'status': 'error', 'message': fk_error, 'data': data})
            continue

        # Duplicate check
        key = tuple(data.get(f, '').strip().lower() for f in dup_criteria)
        if key in existing or key in seen:
            rows_out.append({'status': 'duplicate', 'data': data})
        else:
            seen.add(key)
            rows_out.append({'status': 'ok', 'data': data})

    return rows_out

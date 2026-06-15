import io
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from dashboard.models import (
    Asset, Commodity, Company, Country, Currency, Policy_Level, Policy_Subcategory,
    Policy_Type, Sector, SubnationalRegion, SubSector,
)
from .constants import SHEET_COLUMNS

_HEADER_FILL = PatternFill(start_color='1F7A4A', end_color='1F7A4A', fill_type='solid')
_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_ALIGN = Alignment(horizontal='center')


def build_template():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, columns in SHEET_COLUMNS.items():
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = max(len(col_name) + 4, 15)

    _build_reference_sheet(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_reference_sheet(wb):
    ws = wb.create_sheet('_Référence')
    bold = Font(bold=True)
    sections = [
        ('Countries', Country.objects.values_list('name', flat=True)),
        ('SubnationalRegions', SubnationalRegion.objects.values_list('name', flat=True)),
        ('Commodities', Commodity.objects.values_list('name', flat=True)),
        ('Policy_Types', Policy_Type.objects.values_list('name', flat=True)),
        ('Policy_Subcategories', Policy_Subcategory.objects.values_list('name', flat=True)),
        ('Policy_Levels', Policy_Level.objects.values_list('name', flat=True)),
        ('Currencies', Currency.objects.values_list('code', flat=True)),
        ('Sectors', Sector.objects.values_list('name', flat=True)),
        ('SubSectors', SubSector.objects.values_list('name', flat=True)),
        ('Companies', Company.objects.values_list('name', flat=True)),
        ('Assets', Asset.objects.values_list('name', flat=True)),
    ]
    row = 1
    for section_name, qs in sections:
        ws.cell(row=row, column=1, value=section_name).font = bold
        row += 1
        for name in qs:
            ws.cell(row=row, column=1, value=name)
            row += 1
        row += 1

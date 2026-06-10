import io
import openpyxl
from django.test import TestCase
from dashboard.models import Country, SubnationalRegion, Policy_Type, Policy_Subcategory
from imports.services.excel_parser import parse_file
from imports.services.constants import SHEET_COLUMNS


def _make_xlsx(sheet_data):
    """Build an in-memory .xlsx from {sheet_name: [[header…], [row…], …]}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(r_idx, c_idx, val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ParserCountryTest(TestCase):
    def test_valid_row_is_ok(self):
        buf = _make_xlsx({'Country': [
            ['name', 'water_ownership', 'land_ownership', 'water_Governance', 'land_Governance'],
            ['France', 'public', 'private', '', ''],
        ]})
        result = parse_file(buf)
        self.assertEqual(result['Country'][0]['status'], 'ok')
        self.assertEqual(result['Country'][0]['data']['name'], 'France')

    def test_missing_required_field_is_error(self):
        buf = _make_xlsx({'Country': [
            ['name', 'water_ownership', 'land_ownership', 'water_Governance', 'land_Governance'],
            ['', 'public', 'private', '', ''],
        ]})
        result = parse_file(buf)
        self.assertEqual(result['Country'][0]['status'], 'error')
        self.assertIn('name', result['Country'][0]['message'])

    def test_existing_db_record_is_duplicate(self):
        Country.objects.create(name='France', water_ownership='pub', land_ownership='priv')
        buf = _make_xlsx({'Country': [
            ['name', 'water_ownership', 'land_ownership', 'water_Governance', 'land_Governance'],
            ['France', 'public', 'private', '', ''],
        ]})
        result = parse_file(buf)
        self.assertEqual(result['Country'][0]['status'], 'duplicate')

    def test_case_insensitive_duplicate_detection(self):
        Country.objects.create(name='France', water_ownership='pub', land_ownership='priv')
        buf = _make_xlsx({'Country': [
            ['name', 'water_ownership', 'land_ownership', 'water_Governance', 'land_Governance'],
            ['FRANCE', 'public', 'private', '', ''],
        ]})
        result = parse_file(buf)
        self.assertEqual(result['Country'][0]['status'], 'duplicate')

    def test_within_file_duplicate_is_duplicate(self):
        buf = _make_xlsx({'Country': [
            ['name', 'water_ownership', 'land_ownership', 'water_Governance', 'land_Governance'],
            ['France', 'public', 'private', '', ''],
            ['France', 'public', 'private', '', ''],
        ]})
        result = parse_file(buf)
        self.assertEqual(result['Country'][0]['status'], 'ok')
        self.assertEqual(result['Country'][1]['status'], 'duplicate')

    def test_empty_rows_are_skipped(self):
        buf = _make_xlsx({'Country': [
            ['name', 'water_ownership', 'land_ownership', 'water_Governance', 'land_Governance'],
            ['', '', '', '', ''],
        ]})
        result = parse_file(buf)
        self.assertEqual(result['Country'], [])


class ParserFKTest(TestCase):
    def test_invalid_fk_is_error(self):
        buf = _make_xlsx({'SubnationalRegion': [
            ['name', 'description', 'country_name'],
            ['Bretagne', '', 'NonExistent'],
        ]})
        result = parse_file(buf)
        self.assertEqual(result['SubnationalRegion'][0]['status'], 'error')
        self.assertIn('country_name', result['SubnationalRegion'][0]['message'])

    def test_valid_fk_from_db_is_ok(self):
        Country.objects.create(name='France', water_ownership='pub', land_ownership='priv')
        buf = _make_xlsx({'SubnationalRegion': [
            ['name', 'description', 'country_name'],
            ['Bretagne', '', 'France'],
        ]})
        result = parse_file(buf)
        self.assertEqual(result['SubnationalRegion'][0]['status'], 'ok')

    def test_fk_resolved_from_same_file(self):
        """SubnationalRegion can reference a Country defined in the same file."""
        buf = _make_xlsx({
            'Country': [
                ['name', 'water_ownership', 'land_ownership', 'water_Governance', 'land_Governance'],
                ['NewCountry', 'pub', 'priv', '', ''],
            ],
            'SubnationalRegion': [
                ['name', 'description', 'country_name'],
                ['NewRegion', '', 'NewCountry'],
            ],
        })
        result = parse_file(buf)
        self.assertEqual(result['Country'][0]['status'], 'ok')
        self.assertEqual(result['SubnationalRegion'][0]['status'], 'ok')

    def test_policy_level_subcategory_fk(self):
        pt = Policy_Type.objects.create(name='TypeA')
        Policy_Subcategory.objects.create(name='SubA', policy_type=pt)
        cols = SHEET_COLUMNS['Policy_Level']
        values = {col: '' for col in cols}
        values.update({'name': 'Level1', 'score': '3.0', 'subcategory_name': 'SubA', 'policy_type_name': 'TypeA'})
        buf = _make_xlsx({'Policy_Level': [cols, [values[c] for c in cols]]})
        result = parse_file(buf)
        self.assertEqual(result['Policy_Level'][0]['status'], 'ok')

import io
import openpyxl
from django.test import TestCase
from imports.services.constants import SHEET_COLUMNS
from imports.services.excel_template import build_template


class BuildTemplateTest(TestCase):
    def setUp(self):
        self.buffer = build_template()
        self.wb = openpyxl.load_workbook(self.buffer)

    def test_all_data_sheets_present(self):
        for sheet_name in SHEET_COLUMNS:
            self.assertIn(sheet_name, self.wb.sheetnames, f'Missing sheet: {sheet_name}')

    def test_reference_sheet_present(self):
        self.assertIn('_Référence', self.wb.sheetnames)

    def test_sheet_headers_match_constants(self):
        for sheet_name, expected_cols in SHEET_COLUMNS.items():
            ws = self.wb[sheet_name]
            actual_headers = [ws.cell(1, i + 1).value for i in range(len(expected_cols))]
            self.assertEqual(actual_headers, expected_cols, f'Wrong headers in {sheet_name}')

    def test_returns_bytes_buffer(self):
        buf = build_template()
        self.assertIsInstance(buf, io.BytesIO)
        self.assertGreater(len(buf.getvalue()), 0)
